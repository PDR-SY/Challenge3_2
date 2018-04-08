# -*- coding: utf-8 -*-

from openpyxl import load_workbook # ????????????
from openpyxl import Workbook # ????????????
from datetime import datetime # ?????????????

def combine():
    excel = load_workbook('courses.xlsx')
    student = excel.get_sheet_by_name('students')
    time = excel.get_sheet_by_name('time')

    dict_stu = {}
    dict_time = {}
    for i in range(2,student.max_row+1):
        dict_stu[student.cell(row=i,column=2).value] = {'create_time':student.cell(row=i,column=1).value,'counts':student.cell(row=i,column=3).value}

    for i in range(2,time.max_row+1):
        dict_time[time.cell(row=i,column=2).value] = {'create_time':time.cell(row=i,column=1).value,'time':time.cell(row=i,column=3).value}


    combine = excel.create_sheet('combine',0)
    combine.cell(row=1,column=1).value = '创建时间'
    combine.cell(row=1,column=2).value = '课程名称'
    combine.cell(row=1,column=3).value = '学习人数'
    combine.cell(row=1,column=4).value = '学习时间'
    iRow = 1
    for key in dict_stu.keys():
        iRow+=1
        combine.cell(row =iRow,column = 1).value = datetime.strftime(dict_stu[key]['create_time'],'%Y-%m-%d %H:%M:%S')
        combine.cell(row =iRow,column = 2).value = key
        combine.cell(row =iRow,column = 3).value = dict_stu[key]['counts']
        combine.cell(row =iRow,column = 4).value = dict_time[key]['time']
    excel.save('courses.xlsx')

def split():
    excel = load_workbook('courses.xlsx')
    excel.create_sheet('combine',0)
    combine = excel.get_sheet_by_name('combine')
    dict_combine = {}
    for i in range(2,combine.max_row+1):
        da_time = datetime.strptime(str(combine.cell(row=i,column=1).value),'%Y-%m-%d %H:%M:%S')
        dict_combine.setdefault(da_time.year,[])

        list_combine =  dict_combine[da_time.year]
        list_combine.append({'create_time':combine.cell(row=i,column=1).value,
        'name':combine.cell(row=i,column=2).value,'counts':combine.cell(row=i,column=3).value,
        'times':combine.cell(row=i,column=4).value})
        dict_combine[da_time.year] = list_combine

    
    
    for key in dict_combine.keys():
        new_table = Workbook()
        year = new_table.create_sheet(str(key),0)
        year.cell(row=1,column=1).value = '创建时间'
        year.cell(row=1,column=2).value = '课程名称'
        year.cell(row=1,column=3).value = '学习人数'
        year.cell(row=1,column=4).value = '学习时间'
        for x in range(0,len(dict_combine[key])-1):
            print(dict_combine[key][x])
            year.cell(row=x+2,column=2).value = dict_combine[key][x]['name']
            year.cell(row=x+2,column=1).value = datetime.strptime(dict_combine[key][x]['create_time'],'%Y-%m-%d %H:%M:%S')
            year.cell(row=x+2,column=3).value = dict_combine[key][x]['counts']
            year.cell(row=x+2,column=4).value = dict_combine[key][x]['times']
        new_table.save(str(key)+'.xlsx')
# ??
if __name__ == '__main__':
    combine()
    split()
